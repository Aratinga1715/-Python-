#!/usr/bin/env python
# coding: utf-8

# In[3]:


import os
import shutil
import sys
import re
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import random


# In[4]:


#Отсчет от 0
count = 0
#C:\VKR\data.txt


# In[5]:


user_filename = input("Введите путь к файлу: ")
with open(user_filename, 'r', encoding='utf-8') as file:
    content = file.read()
    print(content)


# In[6]:


with open(user_filename, 'r', encoding='utf-8') as file:
    content = file.read()
    print(content)


# In[7]:


start_word = 'Руководители:'
end_symbol = ' Сотрудники:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]
        for surname in paragraph.split(';'):
          print(surname.title())
    else:
        print("Абзац с заданным концом не найден в файле.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[8]:


start_word = 'Сотрудники:'
end_symbol = ' Организации:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.title())
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[9]:


with open(user_filename, 'r', encoding='utf-8') as file:
 content = file.read()
start_word = 'Организации:'
end_symbol = ' Номера:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.strip())
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[10]:


with open(user_filename, 'r', encoding='utf-8') as file:
 content = file.read()
start_word = 'Номера: '
end_symbol = ' Количество мероприятий план:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.strip())
            
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[11]:


with open(user_filename, 'r', encoding='utf-8') as file:
 content = file.read()
start_word = 'Количество мероприятий план:'
end_symbol = ' Количество мероприятий факт:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.strip())
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[12]:


with open(user_filename, 'r', encoding='utf-8') as file:
 content = file.read()
start_word = 'Количество мероприятий факт:'
end_symbol = ' Статус по сертификации:'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.strip())
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[13]:


with open(user_filename, 'r', encoding='utf-8') as file:
 content = file.read()
start_word = 'Статус по сертификации:'
end_symbol = '.'

start_index = content.find(start_word)
if start_index != -1:
    end_index = content.find(end_symbol, start_index)
    if end_index != -1:
        paragraph = content[start_index:end_index + 1]  # Including the end symbol
        for surname in paragraph.split(';'):
            print(surname.strip())
    else:
        print("Точка не найдена в заданном абзаце.")
else:
    print("Абзац с заданным началом не найден в файле.")


# In[14]:


# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера', 'Количество мероприятий план', 
               'Количество мероприятий факт:', 'Статус по сертификации:']]

# Создаем словарь для каждой категории
category_data = {category: [] for sublist in categories for category in sublist}

# Чтение данных из файла
with open(user_filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        category_data[current_category].append(line)

# Выводим результат
for category, data in category_data.items():
    for item in data:
        print(item)
    print()


# In[15]:


# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Импортируем данные
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]
data = {category: [] for sublist in categories for category in sublist}
with open(user_filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Обработка естественного языка
stemmer = PorterStemmer()

# Функция для форматирования номера телефона
def format_phone_number(phone_number):
  phone_number = str(phone_number).replace(" ", "").replace("(", "").replace(")", "")
  phone_number = str(phone_number).replace("-", "")
  phone_number = str(phone_number).replace(";", "")
  phone_number = str(phone_number).replace(".", "")
  if phone_number.startswith("8"):
    phone_number = phone_number[1:]
  phone_number = "7" + phone_number
  return phone_number

for category, lines in data.items():
    print(f'{category}:')
    for line in lines:
        # Извлечение номера телефона из строки
        phone_number = re.findall(r'\d+', line)
        # Проверка длины номера
        if phone_number and len(phone_number[0]) == 11:
            phone_number = phone_number[0]  # Берем первый найденный номер
            # Форматирование номера телефона
            phone_number = format_phone_number(phone_number)
            print(phone_number)
    print()


# In[16]:


# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Импортируем данные
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]
data = {category: [] for sublist in categories for category in sublist}
with open(user_filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Обработка естественного языка
stemmer = PorterStemmer()

# Обработка статуса сертификации
def process_status(status_str):
    try:
        status = int(status_str.strip())
        return 1 if status > 1 else status
    except ValueError:
        return status_str  # Возвращаем исходную строку, если преобразование в число невозможно

# Обработка данных по сертификации
for line in data['Статус по сертификации:']:
    status = process_status(line)
    print(status)   


# In[17]:


from collections import Counter

# Категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]

# Данные (извлечены из категорий)
data = {category: [] for sublist in categories for category in sublist}

# Добавьте список организаций сюда
data['Организации'] = [
    "ООО \"Газпром\"",
    "ПАО \"Роснефть\"",
    "ООО \"Яндекс\"",
    "АО \"Лукойл\"",
    "АО \"Авантел\"",
    "ПАО \"Магнитогорский металлургический комбинат\""
]

# Очистка и группировка
cleaned_organizations = []
for org in data['Организации']:
    # Удаляем кавычки
    org = org.replace('"', '')
    # Удаляем пробелы в начале и конце
    org = org.strip()
    # Приводим к нижнему регистру
    org = org.lower()
    # Удаляем все символы кроме букв, цифр и пробелов
    org = re.sub(r'[^\w\s]', '', org)
    cleaned_organizations.append(org)

org_counts = Counter(org.split()[0] for org in cleaned_organizations)

# Вывод результатов
print("Анализ организаций:")
for org_type, count in org_counts.items():
    print(f"Тип организации: {org_type}, Количество: {count}")

# Группировка по полному названию
org_counts = Counter(cleaned_organizations)
for org, count in org_counts.items():
    print(f"Название организации: {org}, Количество: {count}")


# In[18]:


# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение данных из файла
with open(user_filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Обработка естественного языка
stemmer = PorterStemmer()

# Обработка статуса сертификации
def process_status(status_str):
    try:
        status = int(status_str.strip())
        return 1 if status > 1 else status
    except ValueError:
        return status_str  # Возвращаем исходную строку, если преобразование в число невозможно

# Обработка данных по сертификации
certification_values = []
for line in data['Статус по сертификации:']:
    status = process_status(line)
    certification_values.append(status)

# Подсчет количества сертификаций для каждого значения
unique_values, counts = np.unique(certification_values, return_counts=True)

# Построение круговой диаграммы
if len(counts) > 1:
    plt.pie(counts, labels=["Не прошли", "Прошли"], autopct='%1.1f%%', startangle=90)
    plt.axis('equal')  # Округляем диаграмму
    plt.title("Распределение сотрудников по статусу сертификации")
    plt.show()
else:
    print("В данных нет записей о людях, которые прошли сертификацию.")


# In[19]:


# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение данных из файла
with open(user_filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Обработка естественного языка
stemmer = PorterStemmer()

# Извлекаем числа из категорий
plan_events = []
fact_events = []
certification_status = []

for category, values in data.items():
    for value in values:
        # Используем регулярное выражение для извлечения чисел, разделенных точкой с запятой
        matches = re.findall(r'\d+', value)
        if matches:
            if "план" in category:
                plan_events.extend(matches)
            elif "факт" in category:
                fact_events.extend(matches)
            elif "сертификации" in category:
                certification_status.extend(matches)  # Добавляем данные о сертификации

# Форматируем данные для записи в файл
plan_str = ', '.join(str(x) for x in plan_events)
fact_str = ', '.join(str(x) for x in fact_events)
certification_str = ', '.join(str(x) for x in certification_status)

# Записываем данные в файл
with open('aaaaaaa.txt', 'w', encoding='utf-8') as file:
    file.write('Количество мероприятий план:\n')
    file.write(plan_str)
    file.write('\n')

    file.write('Количество мероприятий факт:\n')
    file.write(fact_str)
    file.write('\n')

    file.write('Статус по сертификации:\n')
    file.write(certification_str)

# Печатаем содержимое файла в консоль для проверки кодировки
with open('aaaaaaa.txt', 'r', encoding='utf-8') as file:
    data = file.read()
    print(data)


# In[20]:


# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение данных из файла
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:  # Замените "data.txt" на имя вашего файла
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Обработка естественного языка
stemmer = PorterStemmer()

# Извлекаем числа из категорий
plan_events = []
fact_events = []

for category, values in data.items():
    for value in values:
        # Используем регулярное выражение для извлечения чисел, разделенных точкой с запятой
        matches = re.findall(r'\d+', value)
        if matches:
            if "план" in category:
                plan_events.extend(matches)
            elif "факт" in category:
                fact_events.extend(matches)

# Преобразование строк в числа
plan_events = [int(x) for x in plan_events]
fact_events = [int(x) for x in fact_events]

# Функция для симуляции отклонений
def simulate_deviation(data, deviation_factor=0.1):
    """
    Добавляет случайное отклонение к каждому значению в списке.

    Args:
        data: Список значений для модификации.
        deviation_factor: Максимальный процент отклонения от каждого значения.

    Returns:
        Список значений с добавленным отклонением.
    """
    return [value + random.uniform(-value * deviation_factor, value * deviation_factor) for value in data]

# Применение функции к данным
simulated_plan = simulate_deviation(plan_events)
simulated_fact = simulate_deviation(fact_events)

# Визуализация результатов
plt.figure(figsize=(15, 8))  # Увеличиваем размер фигуры
plt.plot(simulated_plan, label="Симулированный план", marker="o")
plt.plot(simulated_fact, label="Симулированный факт", marker="x")
plt.plot(plan_events, label="Исходный план", linestyle="--")
plt.plot(fact_events, label="Исходный факт", linestyle="--")
plt.xlabel("Мероприятие", fontsize=14)  # Увеличиваем размер шрифта подписи оси X
plt.ylabel("Количество мероприятий", fontsize=14)  # Увеличиваем размер шрифта подписи оси Y
plt.title("Имитационное моделирование количества мероприятий", fontsize=16)  # Увеличиваем размер шрифта заголовка
plt.legend(fontsize=12)  # Увеличиваем размер шрифта легенды
plt.grid(True)  # Добавляем сетку
plt.tight_layout()  # Улучшаем компоновку
plt.show()


# In[26]:


import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import re
import openpyxl

# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение данных из файла
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:  # Замените "data.txt" на имя вашего файла
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Извлекаем числа из категорий
extracted_data = {}
for category, values in data.items():
    extracted_data[category] = []
    for value in values:
        # Используем регулярное выражение для извлечения чисел, разделенных точкой с запятой
        matches = re.findall(r'\d+', value)
        if matches:
            extracted_data[category].extend(matches)

# Создание Excel-файла
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов (категории)
headers = [category for category_list in categories for category in category_list]
worksheet.append(headers)

# Заполнение данных
max_length = max(len(values) for values in extracted_data.values())  # Определяем максимальное количество данных

for i in range(max_length):
    row = []
    for category in headers:
        if i < len(extracted_data.get(category, [])):
            row.append(extracted_data[category][i])
        else:
            row.append(None)
    worksheet.append(row)

# Сохранение файла
workbook.save("data.xlsx")

# Загружаем данные из Excel-файла
df = pd.read_excel("data.xlsx")

# Выводим таблицу
print(df)


# In[34]:


import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import re
import openpyxl

# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение числовых данных из файла aaaaaaa.txt
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Извлечение чисел из категорий
extracted_data = {}
for category, values in data.items():
    extracted_data[category] = []
    for value in values:
        matches = re.findall(r'\d+', value)
        if matches:
            extracted_data[category].extend(matches)

# Чтение текстовых данных из файла data.txt
with open("data.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Извлечение текста из категорий "Руководители", "Сотрудники", "Организации", "Номера"
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                extracted_data[category].append(line.replace(category, '').strip())

# Создание Excel-файла
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов (категории)
headers = [category for category_list in categories for category in category_list]
worksheet.append(headers)

# Заполнение данных
max_length = max(len(values) for values in extracted_data.values())  # Определяем максимальное количество данных

for i in range(max_length):
    row = []
    for category in headers:
        if i < len(extracted_data.get(category, [])):
            row.append(extracted_data[category][i])
        else:
            row.append(None)
    worksheet.append(row)

# Сохранение файла
workbook.save("data.xlsx")


# In[22]:


import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import re
import random
from collections import Counter
from nltk.stem import PorterStemmer
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
import os


# Категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера', 'Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]

class DataAnalysisApp:
    def __init__(self, master):
        self.master = master
        master.title("Анализ данных из текстового файла")

        # Стилизация
        style = ttk.Style()
        style.configure('TButton', padding=10, font=('Arial', 12), relief='flat')
        style.configure('TLabel', font=('Arial', 12))
        style.configure('TEntry', font=('Arial', 12), width=50)
        style.configure('TText', font=('Arial', 12), wrap=tk.WORD)

        # Иконка приложения
        icon_path = r"C:\VKR\icon.gif"  # Используйте r-строку или косые слеши
        print(icon_path)  # Вывод пути в консоль
        try:
            self.icon = Image.open(icon_path)
            self.photo = ImageTk.PhotoImage(self.icon)  # Создайте PhotoImage
            master.iconphoto(False, self.photo)  # Установите иконку
        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл иконки не найден!")
        except (IOError, OSError) as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки иконки: {e}")

        # Фрейм для выбора файла
        self.file_frame = tk.Frame(master)
        self.file_frame.pack(pady=10)

        self.file_label = ttk.Label(self.file_frame, text="Выберите файл:")
        self.file_label.pack(side=tk.LEFT)

        self.file_entry = ttk.Entry(self.file_frame)
        self.file_entry.pack(side=tk.LEFT)

        self.browse_button = ttk.Button(self.file_frame, text="Обзор", command=self.browse_file)
        self.browse_button.pack(side=tk.LEFT)

        # Фрейм для вывода результатов
        self.result_frame = tk.Frame(master)
        self.result_frame.pack(pady=10)

        self.result_label = ttk.Label(self.result_frame, text="Результаты:")
        self.result_label.pack()

        self.result_text = tk.Text(self.result_frame, height=10)
        self.result_text.pack()

        # Кнопка для запуска анализа
        self.analyze_button = ttk.Button(master, text="Анализировать", command=self.analyze_data)
        self.analyze_button.pack(pady=10)

        # Кнопка для отображения графика
        self.plot_button = ttk.Button(master, text="Показать график", command=self.plot_data)
        self.plot_button.pack(pady=10)

        # Кнопка для сохранения результатов
        self.save_button = ttk.Button(master, text="Сохранить результаты", command=self.save_results)
        self.save_button.pack(pady=10)

    def browse_file(self):
        """Открывает диалоговое окно для выбора файла."""
        filename = filedialog.askopenfilename(
            initialdir="/",
            title="Выберите файл",
            filetypes=(("Текстовые файлы", "*.txt"), ("Все файлы", "*.*"))
        )
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, filename)

    def analyze_data(self):
        """Анализирует данные из выбранного файла."""
        filename = self.file_entry.get()
        if not filename:
            messagebox.showerror("Ошибка", "Выберите файл!")
            return

        try:
            with open(filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            data = self.process_data(lines)

            # Анализ организаций
            org_counts = Counter(org.split()[0] for org in data['Организации'])
            self.result_text.delete("1.0", tk.END)
            self.result_text.insert(tk.END, "Анализ организаций:\n")
            for org_type, count in org_counts.items():
                self.result_text.insert(tk.END, f"Тип организации: {org_type}, Количество: {count}\n")

            # Извлечение чисел
            plan_events = [int(x) for x in data['Количество мероприятий план']]
            fact_events = [int(x) for x in data['Количество мероприятий факт:']]

            # Сравнение плановых и фактических данных
            self.result_text.insert(tk.END, "\nСравнение плановых и фактических данных:\n")
            for i, (plan, fact) in enumerate(zip(plan_events, fact_events)):
                self.result_text.insert(tk.END, f"Мероприятие {i+1}: план - {plan}, факт - {fact}\n")

            # Дополнительные функции анализа (пример)
            self.result_text.insert(tk.END, "\nДополнительные функции:\n")
            self.result_text.insert(tk.END, f"Среднее количество плановых мероприятий: {sum(plan_events) / len(plan_events)}\n")
            self.result_text.insert(tk.END, f"Среднее количество фактических мероприятий: {sum(fact_events) / len(fact_events)}\n")

        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден!")

    def process_data(self, lines):
        """Обрабатывает данные из файла."""
        data = {category: [] for sublist in categories for category in sublist}
        cleaned_organizations = []

        # Обработка данных
        current_category = ''
        for line in lines:
            line = line.strip()
            for category_list in categories:
                for category in category_list:
                    if category in line:
                        current_category = category
                        break
            if current_category:
                data[current_category].append(line)

            # Обработка организаций 
            if 'Организации' in current_category:
                org = line.replace('"', '')
                org = org.strip()
                org = org.lower()
                org = re.sub(r'[^\w\s]', '', org)
                cleaned_organizations.append(org)

        # Обработка статуса сертификации
        def process_status(status_str):
            try:
                status = int(status_str.strip())
                return 1 if status > 1 else status
            except ValueError:
                return status_str

        certification_values = []
        for line in data['Статус по сертификации:']:
            status = process_status(line)
            certification_values.append(status)

        # Извлечение чисел (с обработкой несоответствий формата)
        plan_events = []
        fact_events = []
        certification_status = []

        for category, values in data.items():
            for value in values:
                # Извлекаем все числа из строки
                matches = re.findall(r'\d+', value)
                
                # Если в строке есть числа, добавляем их в соответствующий список
                if matches:
                    if "план" in category:
                        plan_events.extend(matches)  # Используем extend для добавления списка чисел
                    elif "факт" in category:
                        fact_events.extend(matches)
                    elif "сертификации" in category:
                        certification_status.extend(matches) 

        # Преобразуем данные в числовые
        plan_events = [int(x) for x in plan_events]
        fact_events = [int(x) for x in fact_events]

        # Сохраняем  в  `data`
        data['Количество мероприятий план'] = plan_events
        data['Количество мероприятий факт:'] = fact_events
        data['Статус по сертификации:'] = certification_status
        data['Организации'] = cleaned_organizations

        # Создание DataFrame
        df = pd.DataFrame(data)

        # Преобразование столбцов в числовые, если возможно
        for column in df.columns:
            if column not in ["Руководители: ", "Сотрудники: ", "Организации: "]:
                if all(isinstance(x, (int, float)) for x in df[column] if x is not None):
                    df[column] = pd.to_numeric(df[column], errors='coerce')

        # Возвращаем DataFrame
        return df

    def plot_data(self):
        """Отображает график."""
        filename = self.file_entry.get()
        if not filename:
            messagebox.showerror("Ошибка", "Выберите файл!")
            return

        try:
            with open(filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            data = self.process_data(lines)

            plan_events = data['Количество мероприятий план']
            fact_events = data['Количество мероприятий факт:']

            # Создание графика
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.plot(plan_events, label="План", marker="o", color="blue")  # Настройка цвета
            ax.plot(fact_events, label="Факт", marker="x", color="red")   # Настройка цвета
            ax.set_xlabel("Мероприятие")
            ax.set_ylabel("Количество мероприятий")
            ax.set_title("Сравнение плановых и фактических данных")  # Настройка заголовка
            ax.legend()

            # Встраивание графика в GUI
            canvas = FigureCanvasTkAgg(fig, master=self.master)
            canvas.draw()
            canvas.get_tk_widget().pack(pady=10)

        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден!")

    def save_results(self):
        """Сохраняет результаты в файл Excel."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Файлы Excel", "*.xlsx"), ("Все файлы", "*.*"))
        )
        if not filename:
            return

        try:
            df = self.process_data(self.result_text.get("1.0", tk.END))  # Получаем DataFrame
            df.to_excel(filename, index=False)  # Сохраняем в Excel
            messagebox.showinfo("Успешно", "Результаты успешно сохранены!")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка сохранения: {e}")

# Запуск приложения
root = tk.Tk()
app = DataAnalysisApp(root)
root.mainloop()


# In[33]:


import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import re
import random
from collections import Counter
from nltk.stem import PorterStemmer
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
import os


# Категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера', 'Количество мероприятий план', 'Количество мероприятий факт:', 'Статус по сертификации:']]

class DataAnalysisApp:
    def __init__(self, master):
        self.master = master
        master.title("Анализ данных из текстового файла")

        # Стилизация
        style = ttk.Style()
        style.configure('TButton', padding=10, font=('Arial', 12), relief='flat')
        style.configure('TLabel', font=('Arial', 12))
        style.configure('TEntry', font=('Arial', 12), width=50)
        style.configure('TText', font=('Arial', 12), wrap=tk.WORD)

        # Иконка приложения
        icon_path = r"C:\VKR\icon.jpg"  # Используйте r-строку или косые слеши
        print(icon_path)  # Вывод пути в консоль
        try:
            self.icon = Image.open(icon_path)
            self.photo = ImageTk.PhotoImage(self.icon)  # Создайте PhotoImage
            master.iconphoto(False, self.photo)  # Установите иконку
        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл иконки не найден!")
        except (IOError, OSError) as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки иконки: {e}")

        # Фрейм для выбора файла
        self.file_frame = tk.Frame(master)
        self.file_frame.pack(pady=10)

        self.file_label = ttk.Label(self.file_frame, text="Выберите файл:")
        self.file_label.pack(side=tk.LEFT)

        self.file_entry = ttk.Entry(self.file_frame)
        self.file_entry.pack(side=tk.LEFT)

        self.browse_button = ttk.Button(self.file_frame, text="Обзор", command=self.browse_file)
        self.browse_button.pack(side=tk.LEFT)

        # Фрейм для вывода результатов
        self.result_frame = tk.Frame(master)
        self.result_frame.pack(pady=10)

        self.result_label = ttk.Label(self.result_frame, text="Результаты:")
        self.result_label.pack()

        self.result_text = tk.Text(self.result_frame, height=10)
        self.result_text.pack()

        # Кнопка для запуска анализа
        self.analyze_button = ttk.Button(master, text="Анализировать", command=self.analyze_data)
        self.analyze_button.pack(pady=10)

        # Кнопка для отображения графика
        self.plot_button = ttk.Button(master, text="Показать график", command=self.plot_data)
        self.plot_button.pack(pady=10)

        # Кнопка для сохранения результатов
        self.save_button = ttk.Button(master, text="Сохранить результаты", command=self.save_results)
        self.save_button.pack(pady=10)

    def browse_file(self):
        """Открывает диалоговое окно для выбора файла."""
        filename = filedialog.askopenfilename(
            initialdir="/",
            title="Выберите файл",
            filetypes=(("Текстовые файлы", "*.txt"), ("Все файлы", "*.*"))
        )
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, filename)

    def analyze_data(self):
        """Анализирует данные из выбранного файла."""
        filename = self.file_entry.get()
        if not filename:
            messagebox.showerror("Ошибка", "Выберите файл!")
            return

        try:
            with open(filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            data = self.process_data(lines)

            # Анализ организаций
            org_counts = Counter(org.split()[0] for org in data['Организации'])
            self.result_text.delete("1.0", tk.END)
            self.result_text.insert(tk.END, "Анализ организаций:\n")
            for org_type, count in org_counts.items():
                self.result_text.insert(tk.END, f"Тип организации: {org_type}, Количество: {count}\n")

            # Извлечение чисел
            plan_events = [int(x) for x in data['Количество мероприятий план']]
            fact_events = [int(x) for x in data['Количество мероприятий факт:']]

            # Сравнение плановых и фактических данных
            self.result_text.insert(tk.END, "\nСравнение плановых и фактических данных:\n")
            for i, (plan, fact) in enumerate(zip(plan_events, fact_events)):
                self.result_text.insert(tk.END, f"Мероприятие {i+1}: план - {plan}, факт - {fact}\n")

            # Дополнительные функции анализа (пример)
            self.result_text.insert(tk.END, "\nДополнительные функции:\n")
            self.result_text.insert(tk.END, f"Среднее количество плановых мероприятий: {sum(plan_events) / len(plan_events)}\n")
            self.result_text.insert(tk.END, f"Среднее количество фактических мероприятий: {sum(fact_events) / len(fact_events)}\n")

        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден!")

    def process_data(self, lines):
        """Обрабатывает данные из файла."""
        data = {category: [] for sublist in categories for category in sublist}
        cleaned_organizations = []

        # Обработка данных 
        current_category = ''
        for line in lines:
            line = line.strip()
            for category_list in categories:
                for category in category_list:
                    if category in line:
                        current_category = category
                        break
            if current_category:
                data[current_category].append(line)

            # Обработка организаций 
            if 'Организации' in current_category:
                org = line.replace('"', '')
                org = org.strip()
                org = org.lower()
                org = re.sub(r'[^\w\s]', '', org)
                cleaned_organizations.append(org)

        # Обработка статуса сертификации
        def process_status(status_str):
            try:
                status = int(status_str.strip())
                return 1 if status > 1 else status
            except ValueError:
                return status_str

        certification_values = []
        for line in data['Статус по сертификации:']:
            status = process_status(line)
            certification_values.append(status)

        # Извлечение чисел
        plan_events = []
        fact_events = []
        certification_status = []

        for category, values in data.items():
            for value in values:
                matches = re.findall(r'\d+', value)
                if matches:
                    if "план" in category:
                        plan_events.extend(matches)
                    elif "факт" in category:
                        fact_events.extend(matches)
                    elif "сертификации" in category:
                        certification_status.extend(matches) 

        # Форматируем данные для записи в файл
        plan_str = ', '.join(str(x) for x in plan_events)
        fact_str = ', '.join(str(x) for x in fact_events)
        certification_str = ', '.join(str(x) for x in certification_status)

        # Записываем данные в файл
        with open('aaaaaaa.txt', 'w', encoding='utf-8') as file:
            file.write('Количество мероприятий план:\n')
            file.write(plan_str)
            file.write('\n')

            file.write('Количество мероприятий факт:\n')
            file.write(fact_str)
            file.write('\n')

            file.write('Статус по сертификации:\n')
            file.write(certification_str)

        # Преобразуем данные в числовые
        plan_events = [int(x) for x in plan_events]
        fact_events = [int(x) for x in fact_events]

        # Сохраняем  в  `data`
        data['Количество мероприятий план'] = plan_events
        data['Количество мероприятий факт:'] = fact_events
        data['Статус по сертификации:'] = certification_status
        data['Организации'] = cleaned_organizations

        return data

    def plot_data(self):
        """Отображает график."""
        filename = self.file_entry.get()
        if not filename:
            messagebox.showerror("Ошибка", "Выберите файл!")
            return

        try:
            with open(filename, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            data = self.process_data(lines)

            plan_events = data['Количество мероприятий план']
            fact_events = data['Количество мероприятий факт:']

            # Создание графика
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.plot(plan_events, label="План", marker="o", color="blue")  # Настройка цвета
            ax.plot(fact_events, label="Факт", marker="x", color="red")   # Настройка цвета
            ax.set_xlabel("Мероприятие")
            ax.set_ylabel("Количество мероприятий")
            ax.set_title("Сравнение плановых и фактических данных")  # Настройка заголовка
            ax.legend()

            # Встраивание графика в GUI
            canvas = FigureCanvasTkAgg(fig, master=self.master)
            canvas.draw()
            canvas.get_tk_widget().pack(pady=10)

        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден!")

    def save_results(self):
        """Сохраняет результаты в файл."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=(("Текстовые файлы", "*.txt"), ("Все файлы", "*.*"))
        )
        if not filename:
            return

        try:
            results = self.result_text.get("1.0", tk.END)
            with open(filename, "w", encoding='utf-8') as file:
                file.write(results)
            messagebox.showinfo("Успешно", "Результаты успешно сохранены!")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка сохранения: {e}")

# Запуск приложения
root = tk.Tk()
app = DataAnalysisApp(root)
root.mainloop()


# In[32]:


import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import re
import openpyxl

# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
extracted_data = {category: [] for sublist in categories for category in sublist}

# Чтение числовых данных из файла aaaaaaa.txt
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        extracted_data[current_category].append(line)

# Извлечение чисел из категорий
for category, values in extracted_data.items():
    for i, value in enumerate(values):
        matches = re.findall(r'\d+', value)
        if matches:
            extracted_data[category][i] = matches

# Чтение текстовых данных из файла data.txt
with open("data.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Извлечение текста из категорий "Руководители", "Сотрудники", "Организации", "Номера"
current_category = None
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category and line.startswith(current_category):
        current_category = line.split(':')[0].strip()
        extracted_data[current_category] = []
    elif current_category:
        extracted_data[current_category].append(line.replace(current_category + ':', '').strip())

# Создание Excel-файла
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов (категории)
headers = [category for category_list in categories for category in category_list]
worksheet.append(headers)

# Заполнение данных
max_length = max(len(values) for values in extracted_data.values())  # Определяем максимальное количество данных

for i in range(max_length):
    row = []
    for category in headers:
        if i < len(extracted_data.get(category, [])):
            # Обработка чисел и списков
            if isinstance(extracted_data[category][i], list):
                row.extend(extracted_data[category][i])
            else:
                row.append(extracted_data[category][i])
        else:
            row.append(None)
    worksheet.append(row)

# Сохранение файла
workbook.save("data.xlsx")


# In[49]:


import phonenumbers
import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import re
import openpyxl

# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение числовых данных из файла aaaaaaa.txt
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Извлечение чисел из категорий
extracted_data = {}
for category, values in data.items():
    extracted_data[category] = []
    for value in values:
        matches = re.findall(r'\d+', value)
        if matches:
            extracted_data[category].extend(matches)

# Чтение текстовых данных из файла data.txt
with open("data.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Извлечение текста из категорий "Руководители", "Сотрудники", "Организации", "Номера"
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                extracted_data[category].append(line.replace(category, '').strip())

# Обработка данных в столбцах "Руководители", "Сотрудники", "Организации"
for category in ["Руководители", "Сотрудники", "Организации"]:
    new_data = []
    for value in extracted_data[category]:
        # Разделяем текст по пробелам, запятым и точкам с запятой
        parts = re.split(r'[,\s;]+', value)
        new_data.extend(parts)
    extracted_data[category] = new_data

# Обработка данных в столбце "Номера"
for category in ["Номера"]:
    new_data = []
    for value in extracted_data[category]:
        # Извлекаем номера из строки
        matches = re.findall(r'\d{10,13}', value)
        new_data.extend(matches)
    extracted_data[category] = new_data


# Создание Excel-файла
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов (категории)
headers = [category for category_list in categories for category in category_list]
worksheet.append(headers)

# Заполнение данных
max_length = max(len(values) for values in extracted_data.values())  # Определяем максимальное количество данных

for i in range(max_length):
    row = []
    for category in headers:
        if i < len(extracted_data.get(category, [])):
            row.append(extracted_data[category][i])
        else:
            row.append(None)
    worksheet.append(row)

# Сохранение файла
workbook.save("data.xlsx")


# In[ ]:





# In[50]:


import nltk
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
import re
import openpyxl

# Загружаем стоп-слова
nltk.download('stopwords')
stop_words = set(stopwords.words('russian'))

# Определяем категории
categories = [['Руководители', 'Сотрудники', 'Организации', 'Номера', 'Даты'],
              ['Количество мероприятий план', 'Количество мероприятий факт:'],
              ['Статус по сертификации:']]

# Создаем словарь для каждой категории
data = {category: [] for sublist in categories for category in sublist}

# Чтение числовых данных из файла aaaaaaa.txt
with open("aaaaaaa.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Перебираем строки и разделяем данные по категориям
current_category = ''
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                current_category = category
                break
    if current_category:
        data[current_category].append(line)

# Извлечение чисел из категорий
extracted_data = {}
for category, values in data.items():
    extracted_data[category] = []
    for value in values:
        matches = re.findall(r'\d+', value)
        if matches:
            extracted_data[category].extend(matches)

# Чтение текстовых данных из файла data.txt
with open("data.txt", 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Извлечение текста из категорий "Руководители", "Сотрудники", "Организации", "Номера"
for line in lines:
    line = line.strip()
    for category_list in categories:
        for category in category_list:
            if category in line:
                extracted_data[category].append(line.replace(category, '').strip())

# Обработка данных в столбцах "Руководители", "Сотрудники", "Организации"
for category in ["Руководители", "Сотрудники", "Организации"]:
    new_data = []
    for value in extracted_data[category]:
        # Разделяем текст по пробелам, запятым и точкам с запятой
        parts = re.split(r'[,\s;]+', value)
        new_data.extend(parts)
    extracted_data[category] = new_data

# Обработка данных в столбце "Номера"
extracted_data["Номера"] = []
for line in lines:
    line = line.strip()
    if "Номера:" in line:
        # Находим номера в строке, используя регулярное выражение
        matches = re.findall(r'\d{10,13}', line)
        if matches:
            extracted_data["Номера"].extend(matches)

# Обработка данных в столбце "Даты"
extracted_data["Даты"] = []
for line in lines:
    line = line.strip()
    # Извлекаем даты из строки, используя регулярное выражение
    matches = re.findall(r'\d{2}.\d{2}.\d{4}', line)
    if matches:
        extracted_data["Даты"].extend(matches)

# Создание Excel-файла
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов (категории)
headers = [category for category_list in categories for category in category_list]
worksheet.append(headers)

# Заполнение данных
max_length = max(len(values) for values in extracted_data.values())  # Определяем максимальное количество данных

for i in range(max_length):
    row = []
    for category in headers:
        if i < len(extracted_data.get(category, [])):
            row.append(extracted_data[category][i])
        else:
            row.append(None)
    worksheet.append(row)

# Сохранение файла
workbook.save("data.xlsx")


# In[ ]:




